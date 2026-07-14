import streamlit as st
import pandas as pd
import numpy as np
import pypdf
import io
import patoolib
import os
import shutil
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

# --- THIẾT LẬP TRANG ---
st.set_page_config(
    page_title="Hệ thống Quản trị Báo cáo Xe & Chi Phí Tự Động",
    page_icon="🚚",
    layout="wide"
)

# --- CSS ---
st.markdown("""
<style>
.main-title { font-size: 26px; font-weight: bold; color: #0F172A; margin-bottom: 5px; }
.sub-title { font-size: 14px; color: #475569; margin-bottom: 20px; }
.qt-container { background-color: #F1F5F9; padding: 15px; border-radius: 8px; border-left: 5px solid #2563EB; margin-bottom: 20px; }
.qt-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin-top: 10px; }
.qt-section-title { font-weight: bold; color: #1E3A8A; margin-top: 10px; border-bottom: 2px solid #CBD5E1; padding-bottom: 3px; grid-column: span 2; }
.qt-row { display: flex; justify-content: space-between; padding: 5px 0; border-bottom: 1px dashed #E2E8F0; font-size: 14px; }
.qt-label { font-weight: 500; color: #475569; }
.qt-value { font-weight: bold; color: #0F172A; }
.qt-total-box { grid-column: span 2; background-color: #EFF6FF; padding: 10px; border-radius: 6px; margin-top: 10px; border: 1px solid #BFDBFE; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-title">🚚 HỆ THỐNG QUẢN TRỊ CHI PHÍ VẬN HÀNH & KIỂM SOÁT ĐỊNH MỨC ĐOÀN XE</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">Hợp nhất dữ liệu thô từ file XML Hóa đơn, Bảng kê nhiên liệu thực tế và bảng Chi phí phát sinh ngoại lộ để lập báo cáo.</div>', unsafe_allow_html=True)

# =====================================================================
# 1. HÀM TIỆN ÍCH
# =====================================================================
def normalize_text(text):
    text = str(text).lower().strip()
    text = re.sub(r'[àáạảãâầấậẩẫăằắặẳẵ]', 'a', text)
    text = re.sub(r'[èéẹẻẽêềếệểễ]', 'e', text)
    text = re.sub(r'[ìíịỉĩ]', 'i', text)
    text = re.sub(r'[òóọỏõôồốộổỗơờớợởỡ]', 'o', text)
    text = re.sub(r'[ùúụủũưừứựửữ]', 'u', text)
    text = re.sub(r'[ỳýỵỷỹ]', 'y', text)
    text = re.sub(r'[đ]', 'd', text)
    return re.sub(r'[^a-z0-9]', '', text)

def standardize_plate(text):
    if pd.isna(text):
        return ""
    return str(text).upper().replace("-", "").replace(".", "").replace(" ", "").strip()

def find_header_row(df, target_keywords):
    for i, row in df.iterrows():
        row_values_str = [normalize_text(str(v)) for v in row.values]
        if any(k in r for r in row_values_str for k in target_keywords):
            return i
    return 0

# =====================================================================
# 2. HÀM ĐỊNH DẠNG EXCEL CHUYÊN NGHIỆP
# =====================================================================
def format_excel_sheet(ws, title, df_headers, start_row=0, merge_title=True):
    """
    Định dạng sheet Excel: tiêu đề, border, căn chỉnh, màu sắc.
    """
    # Font và border cơ bản
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    title_font = Font(name='Arial', size=14, bold=True, color='1F4E79')
    content_font = Font(name='Arial', size=10)
    currency_format = '#,##0 "₫"'
    number_format = '#,##0.00'

    # Tiêu đề sheet (nếu có)
    if title and merge_title:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(df_headers))
        cell = ws.cell(row=start_row, column=1, value=title)
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        start_row += 1

    # Header (tên cột)
    for col_idx, col_name in enumerate(df_headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Định dạng các cột (tự động chiều rộng, căn chỉnh)
    for col_idx in range(1, len(df_headers) + 1):
        col_letter = get_column_letter(col_idx)
        # Chiều rộng cột dựa trên nội dung dài nhất
        max_len = 15
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    # Định dạng dữ liệu (căn trái cho text, phải cho số)
    for row in ws.iter_rows(min_row=start_row + 1, max_row=ws.max_row, min_col=1, max_col=len(df_headers)):
        for cell in row:
            cell.border = thin_border
            cell.font = content_font
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right', vertical='center')
                # Định dạng số tiền (nếu tên cột chứa 'VNĐ', 'đ', 'Tiền'...)
                col_name = df_headers[cell.column - 1]
                if any(k in col_name for k in ['VNĐ', 'đ', 'Tiền', 'Thành', 'Giá trị', 'TỔNG', 'CÒN']):
                    cell.number_format = currency_format
                elif 'Km' in col_name or 'Lít' in col_name:
                    cell.number_format = number_format
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Dòng tổng cộng (nếu có) bôi đậm
    last_row = ws.max_row
    if last_row > start_row + 1:
        for col_idx in range(1, len(df_headers) + 1):
            cell = ws.cell(row=last_row, column=col_idx)
            if cell.value and 'TỔNG' in str(cell.value):
                cell.font = Font(name='Arial', size=10, bold=True, color='000000')
                cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            # Nếu là dòng tổng số thì bôi màu nền
            if 'TỔNG CỘNG' in str(cell.value):
                cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

# =====================================================================
# 3. GIẢI NÉN VÀ XÂY DỰNG CACHE TỪ PDF
# =====================================================================
def extract_and_build_cache(uploaded_file, temp_dir="temp_invoice_extract"):
    invoice_cache = {}
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir, ignore_errors=True)
    os.makedirs(temp_dir, exist_ok=True)

    try:
        main_path = os.path.join(temp_dir, uploaded_file.name)
        with open(main_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        try:
            patoolib.extract_archive(main_path, outdir=temp_dir, verbosity=-1)
        except Exception as e:
            st.error(f"Lỗi giải nén tệp đầu vào: {e}")
        if os.path.exists(main_path):
            os.remove(main_path)

        def recursive_extract(folder_path):
            changed = False
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    if file.lower().endswith(('.zip', '.rar')):
                        try:
                            patoolib.extract_archive(file_path, outdir=root, verbosity=-1)
                            os.remove(file_path)
                            changed = True
                        except Exception:
                            pass
            if changed:
                recursive_extract(folder_path)

        recursive_extract(temp_dir)

        # Quét XML
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                if file.lower().endswith('.xml'):
                    try:
                        full_path = os.path.join(root, file)
                        with open(full_path, 'r', encoding='utf-8', errors='ignore') as f:
                            xml_text = f.read()
                        xml_text_upper = xml_text.upper()
                        clean_text = xml_text_upper.replace(" ", "").replace("-", "").replace("_", "")

                        htt_match = re.search(r'<HTTTOAN>([^<]+)</HTTTOAN>', xml_text_upper)
                        htt_value = htt_match.group(1).strip() if htt_match else ""

                        normalized_xml = normalize_text(xml_text)
                        pvoil_keywords = ['pvoil', 'mathe', 'thepvoil', 'pvoileasy', 'theb2b', 'quathe', 'doithe', 'theeasy', 'chuyenkhoan']
                        has_pvoil_keyword = any(k in normalized_xml or k.replace(" ", "") in clean_text for k in pvoil_keywords)

                        if htt_value == "TM":
                            payment_method = "Tiền mặt"
                        elif htt_value == "CK":
                            payment_method = "Thẻ PVOIL"
                        else:
                            payment_method = "Thẻ PVOIL" if has_pvoil_keyword else "Tiền mặt"

                        shd_match = re.search(r'<SHDON>(\d+)</SHDON>', xml_text_upper)
                        if not shd_match:
                            all_numbers = re.findall(r'\d+', xml_text_upper)
                            if all_numbers:
                                shd_match = max(all_numbers, key=len)
                                shd_match = re.match(r'(\d+)', shd_match) if isinstance(shd_match, str) else None
                        if shd_match:
                            if hasattr(shd_match, 'group'):
                                num_raw = shd_match.group(1) if hasattr(shd_match, 'group') else str(shd_match)
                            else:
                                num_raw = str(shd_match)
                        else:
                            num_raw = re.sub(r'\D', '', file)

                        num_stripped = num_raw.lstrip('0') if num_raw else ""

                        if num_raw:
                            invoice_cache[num_raw] = {'payment': payment_method, 'pdf_bytes': None}
                        if num_stripped:
                            invoice_cache[num_stripped] = {'payment': payment_method, 'pdf_bytes': None}
                        if num_raw and num_raw.isdigit():
                            invoice_cache[int(num_raw)] = {'payment': payment_method, 'pdf_bytes': None}
                    except:
                        pass

        # Quét PDF
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                if file.lower().endswith('.pdf'):
                    try:
                        full_path = os.path.join(root, file)
                        with open(full_path, 'rb') as f:
                            pdf_bytes = f.read()
                        number_groups = re.findall(r'\d+', file)
                        for num_raw in number_groups:
                            if len(num_raw) >= 6 or num_raw in invoice_cache or num_raw.lstrip('0') in invoice_cache:
                                num_stripped = num_raw.lstrip('0') if num_raw else ""
                                if num_raw in invoice_cache:
                                    invoice_cache[num_raw]['pdf_bytes'] = pdf_bytes
                                elif num_stripped in invoice_cache:
                                    invoice_cache[num_stripped]['pdf_bytes'] = pdf_bytes
                                else:
                                    invoice_cache[num_raw] = {'payment': 'Tiền mặt', 'pdf_bytes': pdf_bytes}
                                if num_raw and num_raw.isdigit() and int(num_raw) not in invoice_cache:
                                    invoice_cache[int(num_raw)] = {'payment': 'Tiền mặt', 'pdf_bytes': pdf_bytes}
                        # Đọc nội dung PDF
                        try:
                            reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
                            pdf_text = ""
                            for page in reader.pages:
                                pdf_text += page.extract_text() or ""
                            pdf_text_upper = pdf_text.upper()
                            so_matches = re.findall(r'(?:SỐ|SÔ|SỐ HÓA ĐƠN|SÔ HÓA ĐƠN)[\s:]+(\d+)', pdf_text_upper)
                            for so_num in so_matches:
                                if len(so_num) >= 6:
                                    num_raw = so_num
                                    num_stripped = num_raw.lstrip('0')
                                    if num_raw not in invoice_cache:
                                        invoice_cache[num_raw] = {'payment': 'Tiền mặt', 'pdf_bytes': pdf_bytes}
                                    if num_stripped not in invoice_cache:
                                        invoice_cache[num_stripped] = {'payment': 'Tiền mặt', 'pdf_bytes': pdf_bytes}
                                    if num_raw.isdigit() and int(num_raw) not in invoice_cache:
                                        invoice_cache[int(num_raw)] = {'payment': 'Tiền mặt', 'pdf_bytes': pdf_bytes}
                        except:
                            pass
                    except:
                        pass
    finally:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)
    return invoice_cache

# =====================================================================
# 4. XỬ LÝ DỮ LIỆU KINH DOANH
# =====================================================================
def process_business_logic(file_dm, file_bk, file_cp, invoice_cache, filter_month, config_dm):
    # --- Đọc định mức ---
    df_dm_raw = pd.read_excel(file_dm, skiprows=config_dm['idx_header'])
    df_dm_raw.columns = [str(c).strip() for c in df_dm_raw.columns]
    col_dm_bsx = config_dm['col_bsx']
    df_dm_raw['BienSo_Chuan'] = df_dm_raw[col_dm_bsx].apply(standardize_plate)
    df_dm = df_dm_raw.dropna(subset=['BienSo_Chuan']).query("BienSo_Chuan != ''").copy()
    col_tai_xe = config_dm['col_tx']
    col_km_dau = config_dm['col_kmdau']
    col_km_cuoi = config_dm['col_kmcuoi']
    col_dm_rate = config_dm['col_dm']
    df_dm['Dinh_Muc_Rate'] = pd.to_numeric(df_dm[col_dm_rate].astype(str).str.replace(r'[^\d\.]', '', regex=True), errors='coerce').fillna(0)

    # --- Đọc bảng kê nhiên liệu ---
    df_bk_raw = pd.read_excel(file_bk, header=None, dtype=str)
    idx_bk_header = None
    for i, row in df_bk_raw.iterrows():
        row_str = ' '.join([str(v) for v in row.values])
        if 'Số Hóa Đơn' in row_str and 'Biển Kiểm Soát' in row_str:
            idx_bk_header = i
            break
    if idx_bk_header is None:
        idx_bk_header = find_header_row(df_bk_raw, ['bienkiemsoat', 'bienson', 'soxe'])
    if idx_bk_header is None:
        idx_bk_header = 6

    df_bk = pd.read_excel(file_bk, skiprows=idx_bk_header, dtype=str)
    df_bk.columns = [str(c).strip() for c in df_bk.columns]

    col_bk_bsx = next((c for c in df_bk.columns if 'biển' in normalize_text(str(c)) or 'bien' in normalize_text(str(c)) or 'soxe' in normalize_text(str(c))), df_bk.columns[5])
    df_bk['BienSo_Chuan'] = df_bk[col_bk_bsx].apply(standardize_plate)

    col_ngay = next((c for c in df_bk.columns if 'ngay' in normalize_text(str(c)) and 'giao' in normalize_text(str(c))), df_bk.columns[1])
    df_bk['Ngay_DateTime'] = pd.to_datetime(df_bk[col_ngay], errors='coerce')
    if filter_month != "Tất cả các tháng":
        df_bk = df_bk[df_bk['Ngay_DateTime'].dt.month == int(filter_month)].copy()
    df_bk['Ngay_Giao_Dich'] = df_bk['Ngay_DateTime'].dt.strftime('%d/%m/%Y').fillna(df_bk[col_ngay].astype(str))

    col_lit = next((c for c in df_bk.columns if 'số lượng' in normalize_text(str(c)) or 'lit' in normalize_text(str(c))), df_bk.columns[10])
    col_gia = next((c for c in df_bk.columns if 'giá bán' in normalize_text(str(c)) or 'dongia' in normalize_text(str(c))), df_bk.columns[15])
    col_hd = next((c for c in df_bk.columns if 'số hóa đơn' in normalize_text(str(c))), df_bk.columns[22])
    col_mh = next((c for c in df_bk.columns if 'mặt hàng' in normalize_text(str(c)) or 'ten hang' in normalize_text(str(c))), df_bk.columns[9])

    def clean_parse(series):
        return pd.to_numeric(series.astype(str).str.replace(r'[^\d\.]', '', regex=True), errors='coerce').fillna(0)

    df_bk['So_Luong_Lit'] = clean_parse(df_bk[col_lit])
    df_bk['Don_Gia'] = clean_parse(df_bk[col_gia])
    df_bk['Thanh_Tien'] = df_bk['So_Luong_Lit'] * df_bk['Don_Gia']
    df_bk['So_Hoa_Don_Excel'] = df_bk[col_hd].astype(str).str.strip().replace('.0', '', regex=False)
    df_bk['So_Hoa_Don_Excel'] = df_bk['So_Hoa_Don_Excel'].apply(lambda x: '' if (x == 'nan' or x == '') else x)
    df_bk['Mat_Hang'] = df_bk[col_mh].fillna("Nhiên liệu")

    def lookup_payment(row):
        so_hd_excel = row['So_Hoa_Don_Excel']
        if pd.isna(so_hd_excel) or so_hd_excel == '':
            so_hd_excel = ''
        else:
            so_hd_excel = str(so_hd_excel).strip()

        row_text = normalize_text(" ".join([str(v) for v in row.values]))
        excel_says_pvoil = any(k in row_text for k in ['pvoil', 'the', 'ck', 'chuyenkhoan', 'easy'])

        payment = "Tiền mặt"
        pdf_bytes = None
        final_invoice = so_hd_excel
        status = "OK" if so_hd_excel else "Không có số hóa đơn"

        if so_hd_excel:
            hd_str = so_hd_excel
            hd_stripped = hd_str.lstrip('0')
            matched_key = None

            for key in [hd_str, hd_stripped]:
                if key in invoice_cache:
                    matched_key = key
                    payment = invoice_cache[key]['payment']
                    pdf_bytes = invoice_cache[key]['pdf_bytes']
                    break
                if key.isdigit() and int(key) in invoice_cache:
                    matched_key = str(int(key))
                    payment = invoice_cache[int(key)]['payment']
                    pdf_bytes = invoice_cache[int(key)]['pdf_bytes']
                    break

            if matched_key is None and hd_stripped and len(hd_stripped) >= 4:
                for key in invoice_cache:
                    if isinstance(key, str) and key.endswith(hd_stripped):
                        matched_key = key
                        payment = invoice_cache[key]['payment']
                        pdf_bytes = invoice_cache[key]['pdf_bytes']
                        break
                    if isinstance(key, int) and str(key).endswith(hd_stripped):
                        matched_key = str(key)
                        payment = invoice_cache[int(key)]['payment']
                        pdf_bytes = invoice_cache[int(key)]['pdf_bytes']
                        break

            if matched_key is not None:
                final_invoice = matched_key
                status = "OK (PDF)"
            else:
                status = "Không tìm thấy PDF"
        else:
            status = "Không có số hóa đơn (Excel trống)"
            if excel_says_pvoil:
                payment = "Thẻ PVOIL"

        if payment == "Tiền mặt" and excel_says_pvoil:
            payment = "Thẻ PVOIL"

        return payment, pdf_bytes, final_invoice, status

    results = df_bk.apply(lookup_payment, axis=1)
    df_bk['HTTT'] = [r[0] for r in results]
    df_bk['file_pdf_bytes'] = [r[1] for r in results]
    df_bk['So_Hoa_Don_Thuc'] = [r[2] for r in results]
    df_bk['Trang_Thai'] = [r[3] for r in results]

    df_pvoil = df_bk[df_bk['HTTT'] == 'Thẻ PVOIL'].groupby('BienSo_Chuan')['Thanh_Tien'].sum().reset_index(name='ChiPhi_XangXe_PVOil')
    df_cash = df_bk[df_bk['HTTT'] == 'Tiền mặt'].groupby('BienSo_Chuan')['Thanh_Tien'].sum().reset_index(name='ChiPhi_XangXe_TienMat')
    df_lit_totals = df_bk.groupby('BienSo_Chuan')['So_Luong_Lit'].sum().reset_index(name='Tong_Lit_Do_Trong_Ky')

    # --- Đọc chi phí phát sinh ---
    df_cp_raw = pd.read_excel(file_cp, header=None)
    idx_cp = find_header_row(df_cp_raw, ['bienso', 'bienkiemsoat', 'soxe'])
    header_row_1 = df_cp_raw.iloc[idx_cp].ffill()
    header_row_2 = df_cp_raw.iloc[idx_cp + 1].fillna("")
    combined_headers = [normalize_text(str(h1) + "_" + str(h2)) for h1, h2 in zip(header_row_1, header_row_2)]
    df_cp_clean_data = df_cp_raw.iloc[idx_cp + 2:].copy()
    df_cp_clean_data.columns = combined_headers
    idx_col_bsx_cp = next((i for i, h in enumerate(combined_headers) if 'bienso' in h or 'bienkiemsoat' in h or 'soxe' in h), 0)
    df_cp_clean_data['BienSo_Chuan'] = df_cp_clean_data.iloc[:, idx_col_bsx_cp].apply(standardize_plate)
    df_cp_data = df_cp_clean_data.dropna(subset=['BienSo_Chuan']).query("BienSo_Chuan != ''").copy()

    def get_cost(must_contain, fallback=None):
        for i, col_name in enumerate(combined_headers):
            if all(k in col_name for k in must_contain):
                return pd.to_numeric(df_cp_data.iloc[:, i].astype(str).str.replace(r'[^\d\.]', '', regex=True), errors='coerce').fillna(0)
        if fallback:
            for i, col_name in enumerate(combined_headers):
                if all(k in col_name for k in fallback):
                    return pd.to_numeric(df_cp_data.iloc[:, i].astype(str).str.replace(r'[^\d\.]', '', regex=True), errors='coerce').fillna(0)
        return pd.Series(0, index=df_cp_data.index)

    df_cp_data['CP_NB_CauDuong'] = get_cost(['noibo', 'cauduong'], ['noibo', 'lephi'])
    df_cp_data['CP_NB_BenBai'] = get_cost(['noibo', 'ben'], ['noibo', 'vaoben'])
    df_cp_data['CP_NB_BocXep'] = get_cost(['noibo', 'bocxep'])
    df_cp_data['CP_Tinh_CauDuong'] = get_cost(['tinh', 'cauduong'], ['tinh', 'lephi'])
    df_cp_data['CP_Tinh_BenBai'] = get_cost(['tinh', 'ben'], ['tinh', 'vaoben'])
    df_cp_data['CP_Tinh_BocXep'] = get_cost(['tinh', 'bocxep'])
    df_cp_data['CP_RuaXe'] = get_cost(['ruaxe'], ['rua'])

    df_cp_grouped = df_cp_data.groupby('BienSo_Chuan').agg({
        'CP_NB_CauDuong': 'sum', 'CP_NB_BenBai': 'sum', 'CP_NB_BocXep': 'sum',
        'CP_Tinh_CauDuong': 'sum', 'CP_Tinh_BenBai': 'sum', 'CP_Tinh_BocXep': 'sum', 'CP_RuaXe': 'sum'
    }).reset_index()

    # --- Master Merge ---
    df_main = df_dm[['BienSo_Chuan', col_dm_bsx, col_tai_xe, col_km_dau, col_km_cuoi, 'Dinh_Muc_Rate']].copy()
    if config_dm['use_ton_kho']:
        df_main['Ton_Dau_Goc'] = pd.to_numeric(df_dm[config_dm['col_tondau']].astype(str).str.replace(r'[^\d\.]', '', regex=True), errors='coerce').fillna(0)
        df_main['Ton_Cuoi_Goc'] = pd.to_numeric(df_dm[config_dm['col_toncuoi']].astype(str).str.replace(r'[^\d\.]', '', regex=True), errors='coerce').fillna(0)
    else:
        df_main['Ton_Dau_Goc'] = 0
        df_main['Ton_Cuoi_Goc'] = 0

    df_main.rename(columns={col_dm_bsx: 'Bien_So', col_tai_xe: 'Tai_Xe'}, inplace=True)
    df_main = pd.merge(df_main, df_pvoil, on='BienSo_Chuan', how='left').fillna(0)
    df_main = pd.merge(df_main, df_cash, on='BienSo_Chuan', how='left').fillna(0)
    df_main = pd.merge(df_main, df_lit_totals, on='BienSo_Chuan', how='left').fillna(0)
    df_main = pd.merge(df_main, df_cp_grouped, on='BienSo_Chuan', how='left').fillna(0)

    df_main['Km_Dau'] = pd.to_numeric(df_main[col_km_dau].astype(str).str.replace(r'[^\d\.]', '', regex=True), errors='coerce').fillna(0)
    df_main['Km_Cuoi'] = pd.to_numeric(df_main[col_km_cuoi].astype(str).str.replace(r'[^\d\.]', '', regex=True), errors='coerce').fillna(0)
    df_main['Km_ThucTe'] = np.where(df_main['Km_Cuoi'] >= df_main['Km_Dau'], df_main['Km_Cuoi'] - df_main['Km_Dau'], 0)
    df_main['Ton_Dau'] = df_main['Ton_Dau_Goc']
    df_main['Ton_Cuoi'] = df_main['Ton_Cuoi_Goc']
    if config_dm['use_ton_kho']:
        df_main['Nhien_Lieu_Tieu_Thu_Thuc_Te'] = df_main['Ton_Dau'] + df_main['Tong_Lit_Do_Trong_Ky'] - df_main['Ton_Cuoi']
    else:
        df_main['Nhien_Lieu_Tieu_Thu_Thuc_Te'] = df_main['Tong_Lit_Do_Trong_Ky']
    df_main['Nhien_Lieu_Tieu_Thu_Thuc_Te'] = np.where(df_main['Nhien_Lieu_Tieu_Thu_Thuc_Te'] < 0, 0, df_main['Nhien_Lieu_Tieu_Thu_Thuc_Te'])
    df_main['Lit_Dinh_Muc_Cho_Phep'] = df_main['Km_ThucTe'] * (df_main['Dinh_Muc_Rate'] / 100)
    df_main['Lit_Vuot_Dinh_Muc'] = df_main['Nhien_Lieu_Tieu_Thu_Thuc_Te'] - df_main['Lit_Dinh_Muc_Cho_Phep']
    df_main['Lit_Vuot_Dinh_Muc'] = np.where(df_main['Lit_Vuot_Dinh_Muc'] > 0, df_main['Lit_Vuot_Dinh_Muc'], 0)

    df_main['Tong_Cong'] = (df_main['ChiPhi_XangXe_PVOil'] + df_main['ChiPhi_XangXe_TienMat'] +
                            df_main['CP_NB_CauDuong'] + df_main['CP_NB_BenBai'] + df_main['CP_NB_BocXep'] +
                            df_main['CP_Tinh_CauDuong'] + df_main['CP_Tinh_BenBai'] + df_main['CP_Tinh_BocXep'] +
                            df_main['CP_RuaXe'])
    df_main['Con_Phai_Thanh_Toan'] = df_main['Tong_Cong'] - df_main['ChiPhi_XangXe_PVOil']
    return df_main, df_bk

# =====================================================================
# 5. GIAO DIỆN STREAMLIT VỚI SESSION_STATE
# =====================================================================
if 'processed' not in st.session_state:
    st.session_state.processed = False
    st.session_state.df_main = None
    st.session_state.df_bk = None
    st.session_state.df_sys_final = None
    st.session_state.buf_sys = None
    st.session_state.buf_veh = None
    st.session_state.thang_bao_cao = "Tất cả các tháng"

# Sidebar
f_zip = st.sidebar.file_uploader("1. Kho Hóa đơn (Chứa tệp .ZIP hoặc .RAR)", type=["zip", "rar"])
f_dm = st.sidebar.file_uploader("2. File Định mức (.xlsx)", type=["xlsx", "xls"])
f_bk = st.sidebar.file_uploader("3. File Bảng kê NL (.xlsx)", type=["xlsx", "xls"])
f_cp = st.sidebar.file_uploader("4. File CP Phát sinh (.xlsx)", type=["xlsx", "xls"])
thang_bao_cao = st.sidebar.selectbox("Chọn Tháng lập báo cáo:", ["Tất cả các tháng"] + list(range(1, 13)))

config_dm = {}
if f_dm:
    df_dm_headers = pd.read_excel(f_dm, header=None, nrows=10)
    idx_header_dm = 0
    for i, row in df_dm_headers.iterrows():
        row_str = [normalize_text(str(v)) for v in row.values]
        if any('bienso' in r or 'soxe' in r or 'plate' in r for r in row_str):
            idx_header_dm = i
            break
    df_dm_preview = pd.read_excel(f_dm, skiprows=idx_header_dm)
    dm_cols = [str(c) for c in df_dm_preview.columns]
    st.sidebar.markdown("### ⚙️ XÁC NHẬN CỘT FILE ĐỊNH MỨC")
    def guess_idx(cols, keys):
        for i, c in enumerate(cols):
            if any(k in normalize_text(c) for k in keys):
                return i
        return 0

    config_dm['idx_header'] = idx_header_dm
    config_dm['col_bsx'] = st.sidebar.selectbox("Cột Biển số xe:", dm_cols, index=guess_idx(dm_cols, ['bienso', 'soxe', 'plate']))
    config_dm['col_tx'] = st.sidebar.selectbox("Cột Tên lái xe:", dm_cols, index=guess_idx(dm_cols, ['taixe', 'laixe', 'driver']))
    config_dm['col_kmdau'] = st.sidebar.selectbox("Cột Km đầu kỳ:", dm_cols, index=guess_idx(dm_cols, ['kmdau', 'batdau', 'chisocu', 'dauky']))
    config_dm['col_kmcuoi'] = st.sidebar.selectbox("Cột Km cuối kỳ:", dm_cols, index=guess_idx(dm_cols, ['kmcuoi', 'ketthuc', 'chisomoi', 'cuoiky']))
    config_dm['col_dm'] = st.sidebar.selectbox("Cột Định mức khoán (L/100km):", dm_cols, index=guess_idx(dm_cols, ['dinhmuc', 'khoan', 'dm']))
    config_dm['use_ton_kho'] = st.sidebar.checkbox("Báo cáo có tính Tồn Đầu / Tồn Cuối bình dầu", value=False)
    if config_dm['use_ton_kho']:
        config_dm['col_tondau'] = st.sidebar.selectbox("Cột Tồn đầu kỳ:", dm_cols, index=guess_idx(dm_cols, ['tondau', 'dauky']))
        config_dm['col_toncuoi'] = st.sidebar.selectbox("Cột Tồn cuối kỳ:", dm_cols, index=guess_idx(dm_cols, ['toncuoi', 'cuoiky']))

if f_zip and f_dm and f_bk and f_cp:
    if st.sidebar.button("🚀 KÍCH HOẠT QUẢN TRỊ & XUẤT HỆ THỐNG BÁO CÁO", type="primary", use_container_width=True):
        with st.spinner("⏳ Đang giải nén đa tầng và quét bộ dữ liệu XML gốc..."):
            invoice_cache = extract_and_build_cache(f_zip)
        with st.spinner("📊 Đang tổng hợp chi phí và tính toán kiểm soát nhiên liệu..."):
            try:
                df_main, df_bk = process_business_logic(f_dm, f_bk, f_cp, invoice_cache, thang_bao_cao, config_dm)

                # --- BÁO CÁO TỔNG THỂ ---
                df_sys = pd.DataFrame()
                df_sys['STT'] = range(1, len(df_main) + 1)
                df_sys['Biển số xe'] = df_main['Bien_So']
                df_sys['Lái xe'] = df_main['Tai_Xe']
                df_sys['Số Km thực tế'] = df_main['Km_ThucTe']
                df_sys['Định mức khoán (L/100Km)'] = df_main['Dinh_Muc_Rate']
                df_sys['Nhiên liệu định mức (Lít)'] = df_main['Lit_Dinh_Muc_Cho_Phep']
                df_sys['Nhiên liệu tiêu thụ (Lít)'] = df_main['Nhien_Lieu_Tieu_Thu_Thuc_Te']
                df_sys['Dầu vượt định mức (Lít)'] = df_main['Lit_Vuot_Dinh_Muc']
                df_sys['Xăng dầu qua Thẻ PVOil (CK)'] = df_main['ChiPhi_XangXe_PVOil']
                df_sys['Xăng dầu bằng Tiền mặt'] = df_main['ChiPhi_XangXe_TienMat']
                df_sys['Chi Phí Phát Sinh Ngoại Lộ'] = (df_main['Tong_Cong'] - df_main['ChiPhi_XangXe_PVOil'] - df_main['ChiPhi_XangXe_TienMat'])
                df_sys['TỔNG CHI PHÍ VẬN HÀNH'] = df_main['Tong_Cong']
                df_sys['CÒN PHẢI TRẢ TIỀN MẶT'] = df_main['Con_Phai_Thanh_Toan']

                row_total = pd.DataFrame([{
                    'STT': '', 'Biển số xe': 'TỔNG CỘNG ĐOÀN XE', 'Lái xe': '',
                    'Số Km thực tế': df_sys['Số Km thực tế'].sum(),
                    'Định mức khoán (L/100Km)': 0,
                    'Nhiên liệu định mức (Lít)': df_sys['Nhiên liệu định mức (Lít)'].sum(),
                    'Nhiên liệu tiêu thụ (Lít)': df_sys['Nhiên liệu tiêu thụ (Lít)'].sum(),
                    'Dầu vượt định mức (Lít)': df_sys['Dầu vượt định mức (Lít)'].sum(),
                    'Xăng dầu qua Thẻ PVOil (CK)': df_sys['Xăng dầu qua Thẻ PVOil (CK)'].sum(),
                    'Xăng dầu bằng Tiền mặt': df_sys['Xăng dầu bằng Tiền mặt'].sum(),
                    'Chi Phí Phát Sinh Ngoại Lộ': df_sys['Chi Phí Phát Sinh Ngoại Lộ'].sum(),
                    'TỔNG CHI PHÍ VẬN HÀNH': df_sys['TỔNG CHI PHÍ VẬN HÀNH'].sum(),
                    'CÒN PHẢI TRẢ TIỀN MẶT': df_sys['CÒN PHẢI TRẢ TIỀN MẶT'].sum()
                }])
                df_sys_final = pd.concat([df_sys, row_total], ignore_index=True)

                # Xuất Excel tổng thể có định dạng
                buf_sys = io.BytesIO()
                with pd.ExcelWriter(buf_sys, engine='openpyxl') as w_sys:
                    df_sys_final.to_excel(w_sys, index=False, sheet_name="Bao_Cao_Tong_The")
                    ws = w_sys.book["Bao_Cao_Tong_The"]
                    format_excel_sheet(
                        ws,
                        title=f"BÁO CÁO TỔNG HỢP CHI PHÍ VẬN HÀNH ĐOÀN XE - THÁNG {thang_bao_cao}",
                        df_headers=list(df_sys_final.columns),
                        start_row=1
                    )
                buf_sys.seek(0)

                # --- BÁO CÁO CHI TIẾT TỪNG XE ---
                buf_veh = io.BytesIO()
                with pd.ExcelWriter(buf_veh, engine='openpyxl') as w_veh:
                    for _, row in df_main.iterrows():
                        xe_name, tx_name, xe_chuan = row['Bien_So'], row['Tai_Xe'], row['BienSo_Chuan']
                        sh_title = re.sub(r'[\\/*?:\[\]]', '_', str(xe_name))[:30]
                        data_bc1 = [
                            {"STT": 1, "DANH MỤC CHI PHÍ VẬN HÀNH": "Tỷ lệ định mức khoán của xe (Lít/100Km)", "THÔNG TIN CHI TIẾT": f"{row['Dinh_Muc_Rate']:.1f} L/100Km"},
                            {"STT": 2, "DANH MỤC CHI PHÍ VẬN HÀNH": "Số Km đầu kỳ hành trình (Km)", "THÔNG TIN CHI TIẾT": row['Km_Dau']},
                            {"STT": 3, "DANH MỤC CHI PHÍ VẬN HÀNH": "Số Km cuối kỳ hành trình (Km)", "THÔNG TIN CHI TIẾT": row['Km_Cuoi']},
                            {"STT": 4, "DANH MỤC CHI PHÍ VẬN HÀNH": "Số Km chạy thực tế trong tháng (Km)", "THÔNG TIN CHI TIẾT": row['Km_ThucTe']},
                            {"STT": 5, "DANH MỤC CHI PHÍ VẬN HÀNH": "Nhiên liệu tồn đầu kỳ (Lít)", "THÔNG TIN CHI TIẾT": row['Ton_Dau']},
                            {"STT": 6, "DANH MỤC CHI PHÍ VẬN HÀNH": "Nhiên liệu đổ trong kỳ (Lít)", "THÔNG TIN CHI TIẾT": row['Tong_Lit_Do_Trong_Ky']},
                            {"STT": 7, "DANH MỤC CHI PHÍ VẬN HÀNH": "Nhiên liệu tồn cuối kỳ (Lít)", "THÔNG TIN CHI TIẾT": row['Ton_Cuoi']},
                            {"STT": 8, "DANH MỤC CHI PHÍ VẬN HÀNH": "Tổng nhiên liệu tiêu thụ thực tế (Lít)", "THÔNG TIN CHI TIẾT": row['Nhien_Lieu_Tieu_Thu_Thuc_Te']},
                            {"STT": 9, "DANH MỤC CHI PHÍ VẬN HÀNH": "Nhiên liệu định mức cho phép (Lít)", "THÔNG TIN CHI TIẾT": row['Lit_Dinh_Muc_Cho_Phep']},
                            {"STT": 10, "DANH MỤC CHI PHÍ VẬN HÀNH": "⚠️ SỐ LÍT DẦU CHẠY VƯỢT ĐỊNH MỨC KỲ NÀY (LÍT)", "THÔNG TIN CHI TIẾT": row['Lit_Vuot_Dinh_Muc']},
                            {"STT": 11, "DANH MỤC CHI PHÍ VẬN HÀNH": "Chi phí xăng xe mua qua thẻ PVOil (CK)", "THÔNG TIN CHI TIẾT": row['ChiPhi_XangXe_PVOil']},
                            {"STT": 12, "DANH MỤC CHI PHÍ VẬN HÀNH": "Chi phí xăng xe bằng Tiền mặt", "THÔNG TIN CHI TIẾT": row['ChiPhi_XangXe_TienMat']},
                            {"STT": 13, "DANH MỤC CHI PHÍ VẬN HÀNH": "Chi phí chuyển hàng ra bến - Nội bộ", "THÔNG TIN CHI TIẾT": ""},
                            {"STT": "", "DANH MỤC CHI PHÍ VẬN HÀNH": "  - Lệ phí cầu đường", "THÔNG TIN CHI TIẾT": row['CP_NB_CauDuong']},
                            {"STT": "", "DANH MỤC CHI PHÍ VẬN HÀNH": "  - Phí vào bến", "THÔNG TIN CHI TIẾT": row['CP_NB_BenBai']},
                            {"STT": "", "DANH MỤC CHI PHÍ VẬN HÀNH": "  - Phí bốc xếp hàng hóa", "THÔNG TIN CHI TIẾT": row['CP_NB_BocXep']},
                            {"STT": 14, "DANH MỤC CHI PHÍ VẬN HÀNH": "Chi phí chuyển hàng đến đại lý (Tỉnh)", "THÔNG TIN CHI TIẾT": ""},
                            {"STT": "", "DANH MỤC CHI PHÍ VẬN HÀNH": "  - Lệ phí cầu đường", "THÔNG TIN CHI TIẾT": row['CP_Tinh_CauDuong']},
                            {"STT": "", "DANH MỤC CHI PHÍ VẬN HÀNH": "  - Phí vào bến", "THÔNG TIN CHI TIẾT": row['CP_Tinh_BenBai']},
                            {"STT": "", "DANH MỤC CHI PHÍ VẬN HÀNH": "  - Phí bốc xếp hàng hóa", "THÔNG TIN CHI TIẾT": row['CP_Tinh_BocXep']},
                            {"STT": 15, "DANH MỤC CHI PHÍ VẬN HÀNH": "Chi phí rửa xe", "THÔNG TIN CHI TIẾT": row['CP_RuaXe']},
                            {"STT": "★", "DANH MỤC CHI PHÍ VẬN HÀNH": "TỔNG CỘNG CHI PHÍ VẬN HÀNH DOANH NGHIỆP (VNĐ)", "THÔNG TIN CHI TIẾT": row['Tong_Cong']},
                            {"STT": "➔", "DANH MỤC CHI PHÍ VẬN HÀNH": "CÒN PHẢI THANH TOÁN TIỀN MẶT (Tổng - PVOil) (VNĐ)", "THÔNG TIN CHI TIẾT": row['Con_Phai_Thanh_Toan']}
                        ]
                        df_xe_le = df_bk[df_bk['BienSo_Chuan'] == xe_chuan].copy()
                        df_bc2 = pd.DataFrame([{"Thông báo": "Không phát sinh giao dịch nhiên liệu"}])
                        df_bc3 = pd.DataFrame([{"Thông báo": "Không phát sinh hóa đơn điện tử trong kỳ"}])
                        if not df_xe_le.empty:
                            df_bc2 = pd.DataFrame({
                                'Ngày đổ': df_xe_le['Ngay_Giao_Dich'],
                                'Số HD (PDF)': df_xe_le['So_Hoa_Don_Thuc'],
                                'Số HD (Excel)': df_xe_le['So_Hoa_Don_Excel'],
                                'Trạng thái': df_xe_le['Trang_Thai'],
                                'Mặt hàng': df_xe_le['Mat_Hang'],
                                'Số lượng (Lít)': df_xe_le['So_Luong_Lit'],
                                'Thành tiền': df_xe_le['Thanh_Tien']
                            })
                            df_bc3 = pd.DataFrame({
                                'Số Hóa Đơn (PDF)': df_xe_le['So_Hoa_Don_Thuc'],
                                'Số HD (Excel)': df_xe_le['So_Hoa_Don_Excel'],
                                'Trạng thái': df_xe_le['Trang_Thai'],
                                'Ngày phát hành': df_xe_le['Ngay_Giao_Dich'],
                                'Giá trị HD gốc': df_xe_le['Thanh_Tien'],
                                'Hình thức ghi nhận': df_xe_le['HTTT']
                            })

                        # Tạo sheet và ghi dữ liệu
                        ws = w_veh.book.create_sheet(title=sh_title)
                        w_veh.sheets[sh_title] = ws
                        ws.append([f"HỒ SƠ HỢP NHẤT XE: {xe_name} ─ TÀI XẾ PHỤ TRÁCH: {tx_name}"])

                        # Báo cáo 1
                        df_bc1 = pd.DataFrame(data_bc1)
                        df_bc1.to_excel(w_veh, sheet_name=sh_title, startrow=2, index=False)

                        # Báo cáo 2
                        start_bc2 = len(df_bc1) + 5
                        ws.cell(row=start_bc2, column=1, value="📊 BÁO CÁO 2: BẢNG CHI TIẾT CHI PHÍ NGUYÊN LIỆU ĐỔ THỰC TẾ (CÓ TRẠNG THÁI)")
                        df_bc2.to_excel(w_veh, sheet_name=sh_title, startrow=start_bc2+1, index=False)

                        # Báo cáo 3
                        start_bc3 = start_bc2 + len(df_bc2) + 4
                        ws.cell(row=start_bc3, column=1, value="🖨️ BÁO CÁO 3: DANH SÁCH ĐỐI CHIẾU KIỂM KÊ HOÁ ĐƠN XUẤT")
                        df_bc3.to_excel(w_veh, sheet_name=sh_title, startrow=start_bc3+1, index=False)

                        # Định dạng sheet xe
                        # Báo cáo 1
                        format_excel_sheet(
                            ws,
                            title=f"BÁO CÁO CHI TIẾT XE {xe_name} - {tx_name}",
                            df_headers=list(df_bc1.columns),
                            start_row=2
                        )
                        # Báo cáo 2
                        format_excel_sheet(
                            ws,
                            title="BÁO CÁO 2: CHI TIẾT NHIÊN LIỆU",
                            df_headers=list(df_bc2.columns),
                            start_row=start_bc2+1
                        )
                        # Báo cáo 3
                        format_excel_sheet(
                            ws,
                            title="BÁO CÁO 3: KIỂM KÊ HOÁ ĐƠN",
                            df_headers=list(df_bc3.columns),
                            start_row=start_bc3+1
                        )

                    if "Sheet" in w_veh.book.sheetnames:
                        w_veh.book.remove(w_veh.book["Sheet"])
                buf_veh.seek(0)

                st.session_state.processed = True
                st.session_state.df_main = df_main
                st.session_state.df_bk = df_bk
                st.session_state.df_sys_final = df_sys_final
                st.session_state.buf_sys = buf_sys.getvalue()
                st.session_state.buf_veh = buf_veh.getvalue()
                st.session_state.thang_bao_cao = thang_bao_cao
                st.session_state.error = None
            except Exception as e:
                st.session_state.processed = False
                st.session_state.error = str(e)
                st.error(f"❌ Lỗi: {e}")

if st.session_state.processed:
    df_main = st.session_state.df_main
    df_bk = st.session_state.df_bk
    df_sys_final = st.session_state.df_sys_final
    thang_bao_cao = st.session_state.thang_bao_cao

    st.success("🎉 Toàn bộ dữ liệu đoàn xe đã được đồng bộ!")
    b1, b2 = st.columns(2)
    with b1:
        st.download_button(
            "🌐 TẢI BÁO CÁO CHUNG TỔNG THỂ ĐOÀN XE",
            data=st.session_state.buf_sys,
            file_name=f"BC_Chung_Tong_The_{thang_bao_cao}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )
    with b2:
        st.download_button(
            "🗂️ TẢI BỘ BÁO CÁO TÁCH SHEET CHI TIẾT TỪNG XE (CÓ CẢNH BÁO)",
            data=st.session_state.buf_veh,
            file_name=f"Bo_3_Bao_Cao_Tach_Tung_Xe_{thang_bao_cao}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="secondary",
            use_container_width=True
        )

    st.markdown("### 🌐 ĐỀ MỤC 1: BÁO CÁO CHUNG TỔNG THỂ HỆ THỐNG ĐOÀN XE")
    df_display = df_sys_final.copy()
    df_display['Số Km thực tế'] = df_display['Số Km thực tế'].apply(lambda x: f"{x:,.0f} Km" if isinstance(x, (int, float)) else x)
    df_display['Nhiên liệu định mức (Lít)'] = df_display['Nhiên liệu định mức (Lít)'].apply(lambda x: f"{x:,.1f} L" if isinstance(x, (int, float)) else x)
    df_display['Nhiên liệu tiêu thụ (Lít)'] = df_display['Nhiên liệu tiêu thụ (Lít)'].apply(lambda x: f"{x:,.1f} L" if isinstance(x, (int, float)) else x)
    df_display['Dầu vượt định mức (Lít)'] = df_display['Dầu vượt định mức (Lít)'].apply(lambda x: f"{x:,.1f} L" if isinstance(x, (int, float)) else x)
    cols_money = ['Xăng dầu qua Thẻ PVOil (CK)', 'Xăng dầu bằng Tiền mặt', 'Chi Phí Phát Sinh Ngoại Lộ', 'TỔNG CHI PHÍ VẬN HÀNH', 'CÒN PHẢI TRẢ TIỀN MẶT']
    for col in cols_money:
        df_display[col] = df_display[col].apply(lambda x: f"{x:,.0f} đ" if isinstance(x, (int, float)) else x)
    st.dataframe(df_display, use_container_width=True)

    st.markdown("### 🗂️ ĐỀ MỤC 2: BẢNG TRA CỨU CHI TIẾT BỘ 3 BÁO CÁO RIÊNG TỪNG XE (CÓ CẢNH BÁO)")

    for _, row in df_main.iterrows():
        xe_chuan = row['BienSo_Chuan']
        df_xe = df_bk[df_bk['BienSo_Chuan'] == xe_chuan].copy()

        if not df_xe.empty:
            df_combined = pd.DataFrame()
            df_combined['Ngày đổ'] = df_xe['Ngay_Giao_Dich']
            df_combined['Số HD (PDF)'] = df_xe['So_Hoa_Don_Thuc']
            df_combined['Số HD (Excel)'] = df_xe['So_Hoa_Don_Excel']
            df_combined['Trạng thái'] = df_xe['Trang_Thai']
            df_combined['Mặt hàng'] = df_xe['Mat_Hang']
            df_combined['Số lượng (Lít)'] = df_xe['So_Luong_Lit']
            df_combined['Thành tiền (VNĐ)'] = df_xe['Thanh_Tien']
            df_combined['Hình thức TT'] = df_xe['HTTT']
        else:
            df_combined = pd.DataFrame([{"Thông báo": "Không có dữ liệu đổ xăng trong kỳ"}])

        color_bg = '#FEE2E2' if row['Lit_Vuot_Dinh_Muc'] > 0 else '#D1FAE5'
        color_text = '#DC2626' if row['Lit_Vuot_Dinh_Muc'] > 0 else '#16A34A'
        status_text = f"Vượt định mức: {row['Lit_Vuot_Dinh_Muc']:,.1f} Lít dầu tiêu hao ⚠️" if row['Lit_Vuot_Dinh_Muc'] > 0 else "Đạt định mức dầu tiêu chuẩn"

        with st.expander(f"🚗 Xe: {row['Bien_So']} ─ Lái xe: {row['Tai_Xe']} ｜ {status_text}"):
            st.markdown(f"""
<div class="qt-container">
<div class="qt-grid">
<div class="qt-section-title">1. ĐƠN VỊ HÀNH TRÌNH KHAI BÁO CỦA XE</div>
<div class="qt-row" style="grid-column: span 2;"><span class="qt-label">Km Đầu kỳ hành trình:</span><span class="qt-value">{row['Km_Dau']:,.0f} Km</span></div>
<div class="qt-row" style="grid-column: span 2;"><span class="qt-label">Km Cuối kỳ hành trình:</span><span class="qt-value">{row['Km_Cuoi']:,.0f} Km</span></div>
<div class="qt-row" style="grid-column: span 2; background-color: #F8FAFC;"><span class="qt-label">➔ Tổng số Km chạy thực tế trong tháng:</span><span class="qt-value" style="color: #2563EB;">{row['Km_ThucTe']:,.0f} Km</span></div>

<div class="qt-section-title">2. KIỂM SOÁT ĐỊNH MỨC TIÊU HAO NHIÊN LIỆU KỲ NÀY</div>
<div class="qt-row"><span class="qt-label">Tồn đầu kỳ bình xe:</span><span class="qt-value">{row['Ton_Dau']:,.1f} Lít</span></div>
<div class="qt-row"><span class="qt-label">Đổ thêm trong tháng:</span><span class="qt-value">{row['Tong_Lit_Do_Trong_Ky']:,.1f} Lít</span></div>
<div class="qt-row"><span class="qt-label">Tồn cuối kỳ bình xe:</span><span class="qt-value">{row['Ton_Cuoi']:,.1f} Lít</span></div>
<div class="qt-row" style="grid-column: span 2;"><span class="qt-label">⛽ Tỷ lệ định mức khoán tiêu hao:</span><span class="qt-value">{row['Dinh_Muc_Rate']:.1f} Lít/100Km</span></div>
<div class="qt-row" style="grid-column: span 2;"><span class="qt-label">📉 Nhiên liệu định mức công ty cho phép:</span><span class="qt-value">{row['Lit_Dinh_Muc_Cho_Phep']:,.1f} Lít</span></div>
<div class="qt-row" style="grid-column: span 2; background-color: #F1F5F9;"><span class="qt-label">📊 TỔNG TIÊU THỤ THỰC TẾ TRONG KỲ:</span><span class="qt-value">{row['Nhien_Lieu_Tieu_Thu_Thuc_Te']:,.1f} Lít</span></div>
<div class="qt-row" style="grid-column: span 2; background-color: {color_bg};"><span class="qt-label">⚠️ Số lít dầu vượt định mức:</span><span class="qt-value" style="color: {color_text};">{row['Lit_Vuot_Dinh_Muc']:,.1f} Lít</span></div>

<div class="qt-section-title">3. CHI PHÍ NHIÊN LIỆU (PHÂN LOẠI HÓA ĐƠN THỰC TẾ)</div>
<div class="qt-row" style="grid-column: span 2; background-color: #FFFBEB;"><span class="qt-label">💳 Mua qua thẻ PVOil (Công ty chuyển khoản trước) (A1):</span><span class="qt-value" style="color: #D97706;">{row['ChiPhi_XangXe_PVOil']:,.0f} VNĐ</span></div>
<div class="qt-row" style="grid-column: span 2; background-color: #FFFBEB;"><span class="qt-label">💵 Mua bằng Tiền mặt (Lái xe tự trả tại trạm) (A2):</span><span class="qt-value" style="color: #BA2525;">{row['ChiPhi_XangXe_TienMat']:,.0f} VNĐ</span></div>

<div class="qt-section-title">4. CHI PHÍ CHUYỂN HÀNG RA BẾN - NỘI BỘ (B1)</div>
<div class="qt-row" style="grid-column: span 2;"><span class="qt-label"> └─ Lệ phí cầu đường:</span><span class="qt-value">{row['CP_NB_CauDuong']:,.0f} VNĐ</span></div>
<div class="qt-row" style="grid-column: span 2;"><span class="qt-label"> └─ Phí vào bến:</span><span class="qt-value">{row['CP_NB_BenBai']:,.0f} VNĐ</span></div>
<div class="qt-row" style="grid-column: span 2;"><span class="qt-label"> └─ Phí bốc xếp hàng hóa:</span><span class="qt-value">{row['CP_NB_BocXep']:,.0f} VNĐ</span></div>

<div class="qt-section-title">5. CHI PHÍ CHUYỂN HÀNG ĐẾN ĐẠI LÝ - TỈNH (B2)</div>
<div class="qt-row" style="grid-column: span 2;"><span class="qt-label"> └─ Lệ phí cầu đường:</span><span class="qt-value">{row['CP_Tinh_CauDuong']:,.0f} VNĐ</span></div>
<div class="qt-row" style="grid-column: span 2;"><span class="qt-label"> └─ Phí vào bến:</span><span class="qt-value">{row['CP_Tinh_BenBai']:,.0f} VNĐ</span></div>
<div class="qt-row" style="grid-column: span 2;"><span class="qt-label"> └─ Phí bốc xếp hàng hóa:</span><span class="qt-value">{row['CP_Tinh_BocXep']:,.0f} VNĐ</span></div>

<div class="qt-section-title">6. DỊCH VỤ NGOÀI ĐỊNH KỲ (B3)</div>
<div class="qt-row" style="grid-column: span 2;"><span class="qt-label">🧼 Chi phí rửa xe phát sinh:</span><span class="qt-value">{row['CP_RuaXe']:,.0f} VNĐ</span></div>

<div class="qt-total-box">
<div class="qt-row"><span class="qt-label">★ TỔNG CHI PHÍ VẬN HÀNH DOANH NGHIỆP GHI NHẬN:</span><span class="qt-value">{row['Tong_Cong']:,.0f} VNĐ</span></div>
<div class="qt-row" style="border:none; margin-top:5px;"><span class="qt-label" style="color:#16A34A; font-size:15px; font-weight:bold;">➔ SỐ TIỀN MẶT CẦN HOÀN TRẢ QUYẾT TOÁN CHO LÁI XE:</span><span class="qt-value" style="color:#16A34A; font-size:15px; font-weight:bold;">{row['Con_Phai_Thanh_Toan']:,.0f} VNĐ</span></div>
</div>
</div>
</div>
""", unsafe_allow_html=True)

            st.markdown("---")
            st.markdown("#### 📊 BÁO CÁO 2+3: CHI TIẾT NHIÊN LIỆU & ĐỐI CHIẾU HÓA ĐƠN (CÓ CẢNH BÁO)")
            if not df_xe.empty:
                df_display_combined = df_combined.copy()
                if 'Số lượng (Lít)' in df_display_combined.columns:
                    df_display_combined['Số lượng (Lít)'] = df_display_combined['Số lượng (Lít)'].apply(lambda x: f"{x:,.2f}" if isinstance(x, (int, float)) else x)
                if 'Thành tiền (VNĐ)' in df_display_combined.columns:
                    df_display_combined['Thành tiền (VNĐ)'] = df_display_combined['Thành tiền (VNĐ)'].apply(lambda x: f"{x:,.0f}" if isinstance(x, (int, float)) else x)
                st.dataframe(df_display_combined, use_container_width=True)

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_combined.to_excel(writer, index=False, sheet_name="BC_2_3")
                    for sheet in writer.book.worksheets:
                        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
                        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
                        sheet.sheet_properties.pageSetUpPr.fitToPage = True
                        sheet.page_setup.fitToWidth = 1
                        sheet.page_setup.fitToHeight = 0
                output.seek(0)
                st.download_button(
                    label=f"⬇️ TẢI BÁO CÁO 2+3 (Excel) cho xe {row['Bien_So']}",
                    data=output,
                    file_name=f"BC_2_3_Xe_{row['Bien_So']}_{thang_bao_cao}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_excel_{row['BienSo_Chuan']}"
                )
            else:
                st.info("Không có dữ liệu đổ xăng trong kỳ")

            st.markdown("---")
            valid_pdfs = [b for b in df_xe['file_pdf_bytes'] if b is not None]
            if valid_pdfs:
                try:
                    writer = pypdf.PdfWriter()
                    for b in valid_pdfs:
                        reader = pypdf.PdfReader(io.BytesIO(b))
                        if len(reader.pages) > 0:
                            writer.add_page(reader.pages[0])
                    out = io.BytesIO()
                    writer.write(out)
                    st.download_button(
                        f"🖨️ TẢI GỘP {len(valid_pdfs)} PDF HÓA ĐƠN CỦA XE NÀY",
                        out.getvalue(),
                        f"HoaDon_Gop_Xe_{row['BienSo_Chuan']}.pdf",
                        "application/pdf",
                        key=f"dl_{row['BienSo_Chuan']}"
                    )
                except Exception as e:
                    st.error(f"Lỗi nối file PDF: {e}")
            else:
                st.button("🖨️ KHÔNG TÌM THẤY FILE PDF NÀO KHỚP VỚI XE NÀY", disabled=True, key=f"dl_dis_{row['BienSo_Chuan']}")

else:
    if not (f_zip and f_dm and f_bk and f_cp):
        st.info("💡 Vui lòng nạp đầy đủ cả 4 tệp dữ liệu ở cột bên trái.")